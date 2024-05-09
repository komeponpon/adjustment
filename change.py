import openpyxl

#メインファイルの読み込み
wb = openpyxl.load_workbook('直販_納品書_物件別_24年04月_販売L.xlsx')

#値の取得用ファイル
wb2 = openpyxl.load_workbook('240430_2024年度_月別_主要部材原価.xlsx')
ws2 = wb2.active
new_value = ws2['G14'].value #該当金額のセルを選択(ex.400W:G12,415W:G13,430W:G14,PW)

valid_names = [
  #"単400W／BCモジュール　SPR-MAX3-400　　[N-G]",
  #"単400W／BCモジュール　SPR-MAX3-400　　[M2]",
  #"単415Wモジュール　LR5-54HPH-415M",
  "N型単430Wモジュール　SS430M8GFH-18/VNH"
]

#全てのシートに同様の処理
for sheet in wb.sheetnames:
    ws = wb[sheet]
    product_name = ws['C17'].value

    if product_name in valid_names:
        ws['O17'] = new_value
        ws['P17'] = ws['O17'].value * ws['F17'].value

#変更を保存
wb.save('直販_納品書_物件別_24年04月_販売s.xlsx')